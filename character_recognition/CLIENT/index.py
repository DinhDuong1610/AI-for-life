import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import os
import uuid

# Địa chỉ API
url = "http://127.0.0.1:5000/ocr/"

image_path = "image/bangdiem5.jpg"

coordinates = [
    [413,1283,594,1315],
    [606,1283,742,1315],
    [758,1283,922,1315],
    [413,1324,594,1359],
    [604,1322,746,1360],
    [765,1324,922,1359],
    [415,1367,592,1401],
    [602,1364,746,1402],
    [756,1366,924,1402],
    [938,1282,1059,1315],
    [938,1322,1064,1359],
    [940,1366,1062,1399],
    [1077,1537,1247,1568]
]
tolerance = 10
payload = {
    'coordinates[]': [str(coords) for coords in coordinates]
}

# Gửi ảnh và tọa độ tới API
with open(image_path, "rb") as img_file:
    files = {"file": img_file}
    response = requests.post(url, files=files, data=payload)

if response.status_code == 200:
    results = response.json()
    for result in results:
        print(f"Coordinates: {result['coordinates']}")
        print(f"Text: {result['text']}")
        print(f"Confidence: {result['confidence']}")
        print(f"-------------")

    grouped_results = []
    for result in results:
        coords = result['coordinates']
        y = coords[1]  # Lấy min_y
        added_to_group = False

        for group in grouped_results:
            if abs(group[0]['coordinates'][1] - y) <= tolerance:
                group.append(result)
                added_to_group = True
                break

        if not added_to_group:
            grouped_results.append([result])

    for group in grouped_results:
        group.sort(key=lambda item: item['coordinates'][0])

    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    alignment = Alignment(horizontal="center", vertical="center")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    row_idx = 1
    for group in grouped_results:
        col_idx = 1
        for result in group:
            text = result["text"]
            confidence = result["confidence"]
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = alignment

            if confidence < 0.9:
                cell.fill = yellow_fill

            col_idx += 1
        row_idx += 1

    os.makedirs("results_excel", exist_ok=True)

    random_filename = f"OCR_Results_{uuid.uuid4().hex}.xlsx"
    output_file = os.path.join("results_excel", random_filename)

    wb.save(output_file)
    print(f"Data has been saved to {output_file}")

else:
    print("Error:", response.status_code, response.text)

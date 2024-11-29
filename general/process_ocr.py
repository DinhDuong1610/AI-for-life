import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from table_recognition.main import process_image
from character_recognition.intmain import process_image_with_coordinates

# Hàm gom nhóm tọa độ
def group_coordinates(coordinates, threshold):
    grouped = []
    for coord in sorted(coordinates):
        if not grouped or coord - grouped[-1] > threshold:
            grouped.append(coord)
    return grouped

def is_row_valid(row_data, max_col, empty_threshold=0.07):
    filled_cells = sum(1 for cell in row_data if cell)  
    empty_cells = len(row_data) - filled_cells
    return (empty_cells / max_col) < empty_threshold  

# Hàm xử lý ảnh và trả về file Excel
def process_multiple_images_to_excel(image_paths):
    # Tạo Workbook và Worksheet mới
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    # Căn chỉnh dữ liệu trong Excel
    alignment = Alignment(horizontal="center", vertical="center")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    current_row = 1  # Bắt đầu ghi từ dòng đầu tiên

    # Danh sách lưu lại các đường dẫn output
    all_out_paths = []
    title_results = [] 

    for image_path in image_paths:
        ocr_results, out_path ,title_result= process_image_with_coordinates(image_path, process_image(image_path))

        # Thêm đường dẫn kết quả vào danh sách
        all_out_paths.append(out_path)
        
        title_results.append(title_result)

        # Xác định khoảng ngưỡng để gom nhóm hàng/cột
        threshold = 10

        # Lấy tất cả tọa độ X và Y từ kết quả OCR
        x_coords = [result['coordinates'][0] for result in ocr_results]
        y_coords = [result['coordinates'][1] for result in ocr_results]

        # Gom nhóm các tọa độ vào hàng và cột
        x_groups = group_coordinates(x_coords, threshold)
        y_groups = group_coordinates(y_coords, threshold)

        # Xác định số hàng và số cột
        max_row = len(y_groups)
        max_col = len(x_groups)

        # Tạo một ma trận trống để lưu các ô
        matrix = [["" for _ in range(max_col)] for _ in range(max_row)]

        # Chèn các ô OCR vào ma trận
        for result in ocr_results:
            coords = result['coordinates']
            x, y = coords[0], coords[1]

            # Tìm cột gần nhất
            col = min(range(len(x_groups)), key=lambda i: abs(x - x_groups[i]))
            # Tìm hàng gần nhất
            row = min(range(len(y_groups)), key=lambda i: abs(y - y_groups[i]))

            # Ghi dữ liệu vào ma trận
            matrix[row][col] = {
                "text": result["text"],
                "confidence": result["confidence"]
            }

        # Lọc bỏ các hàng không hợp lệ
        valid_matrix = [row for row in matrix if is_row_valid(row, max_col)]

        # Ghi dữ liệu ma trận đã lọc vào Excel
        for row_idx, row_data in enumerate(valid_matrix, start=current_row):
            for col_idx, cell_data in enumerate(row_data, start=1):
                if cell_data:  # Nếu ô có dữ liệu
                    text = cell_data["text"]
                    confidence = cell_data["confidence"]

                    # Nếu là cột B (cột thứ 2), thay thế chữ 'O' bằng '0'
                    if col_idx == 2:  # Cột B (tính từ 1)
                        text = text.replace('O', '0')

                    cell = ws.cell(row=row_idx, column=col_idx, value=text)
                    cell.alignment = alignment


                    if confidence < 0.5:
                        cell.fill = red_fill
                    elif confidence < 0.85:
                        cell.fill = yellow_fill
                    
                else:  # Nếu ô trống
                    ws.cell(row=row_idx, column=col_idx, value="")

        # Thêm khoảng trống giữa các ảnh
        current_row += len(valid_matrix)

    # Tạo thư mục lưu kết quả nếu chưa có
    os.makedirs("results_excel", exist_ok=True)

    # Tạo tên file ngẫu nhiên
    random_filename = f"OCR_Results_{uuid.uuid4().hex}.xlsx"
    output_file = os.path.join("results_excel", random_filename)

    # Lưu workbook vào file
    wb.save(output_file)

    return output_file, random_filename, all_out_paths,title_results

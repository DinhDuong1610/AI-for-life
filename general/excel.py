from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import os
import uuid
from table_recognition.main import process_image
from character_recognition.intmain import process_image_with_coordinates

# Đường dẫn tới ảnh
image_path = "./images/bangdiemcophuong1.png"

# Lấy kết quả từ các hàm nhận diện
ocr_results = process_image_with_coordinates(image_path, process_image(image_path))

# Xác định khoảng ngưỡng để gom nhóm hàng/cột
threshold = 5  # Sai số cho phép (đơn vị pixels)

# Gom nhóm tọa độ X và Y
def group_coordinates(coordinates, threshold):
    grouped = []
    for coord in sorted(coordinates):
        if not grouped or coord - grouped[-1] > threshold:
            grouped.append(coord)
    return grouped

# Lấy tất cả tọa độ X và Y từ kết quả OCR
x_coords = [result['coordinates'][0] for result in ocr_results]
y_coords = [result['coordinates'][1] for result in ocr_results]

# Gom nhóm các tọa độ vào hàng và cột
x_groups = group_coordinates(x_coords, threshold)
y_groups = group_coordinates(y_coords, threshold)

# Xác định số hàng và số cột
max_row = len(y_groups)
max_col = len(x_groups)

# Tạo một ma trận trống để lưu các ô
matrix = [["" for _ in range(max_col)] for _ in range(max_row)]

# Chèn các ô OCR vào ma trận
for result in ocr_results:
    coords = result['coordinates']
    x, y = coords[0], coords[1]

    # Tìm cột gần nhất
    col = min(range(len(x_groups)), key=lambda i: abs(x - x_groups[i]))
    # Tìm hàng gần nhất
    row = min(range(len(y_groups)), key=lambda i: abs(y - y_groups[i]))

    # Ghi dữ liệu vào ma trận
    matrix[row][col] = {
        "text": result["text"],
        "confidence": result["confidence"]
    }

# Tạo Workbook và Worksheet mới
wb = Workbook()
ws = wb.active
ws.title = "OCR Results"

# Căn chỉnh dữ liệu trong Excel
alignment = Alignment(horizontal="center", vertical="center")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Duyệt ma trận và ghi dữ liệu vào Excel
for row_idx, row_data in enumerate(matrix, start=1):
    for col_idx, cell_data in enumerate(row_data, start=1):
        if cell_data:  # Nếu ô có dữ liệu
            text = cell_data["text"]
            confidence = cell_data["confidence"]
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = alignment

            # Nếu độ tin cậy < 0.9, tô màu vàng
            if confidence < 0.9:
                cell.fill = yellow_fill
        else:  # Nếu ô trống
            ws.cell(row=row_idx, column=col_idx, value="")

# Tạo thư mục lưu kết quả nếu chưa có
os.makedirs("results_excel", exist_ok=True)

# Tạo tên file ngẫu nhiên
random_filename = f"OCR_Results_{uuid.uuid4().hex}.xlsx"
output_file = os.path.join("results_excel", random_filename)

# Lưu workbook vào file
wb.save(output_file)
print(f"Data has been saved to {output_file}")
